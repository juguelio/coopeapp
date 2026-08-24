"""Lectura de un cómputo y presupuesto (.xlsx) para armar una foja de medición.

Este módulo NO importa Odoo a propósito: la parte difícil de importar un
cómputo es decidir qué fila es un ítem, y eso se puede probar sin levantar un
servidor. El wizard (`coop.foja.import`) usa estas funciones; los tests las
usan directamente.

Lo que hace y lo que deliberadamente NO hace
--------------------------------------------
Un cómputo trae la estructura que le dio quien lo armó. Sobre el archivo real
de Carriqueo aparecieron, sin buscarlos, cuatro problemas que ningún parser
puede resolver solo:

1. **Doble conteo.** Los rubros de nivel 1 (`1.0 DESARME`, `2.0 ALBAÑILERÍA`)
   traen importe, y sus subítems (`1.1`, `1.2`, …) también. Sumar los dos
   niveles cuenta la obra dos veces. Cuál de los dos es "la foja" depende del
   archivo.
2. **Filas de plantilla sin limpiar.** En una remodelación de oficinas, los
   subítems hablaban de cañerías de agua y estaciones de bombeo cloacal: venían
   del machete y nadie los borró.
3. **Excel se come los códigos.** `5.1` quedó guardado como la fecha
   05/01/2023, y `1.10` colapsó a `1.1`.
4. **Tipos mezclados** en la misma columna: `1.0` como número y `1.10` como
   texto.

Por eso este parser **clasifica y avisa, no decide**. Devuelve las filas con su
nivel, los totales por nivel y la lista de avisos, para que una persona mire y
confirme antes de crear nada. Importar en silencio un cómputo mal leído es
peor que no importarlo: la incidencia de cada ítem es la que después mueve el
% de avance y el aporte de cada socio.
"""

import datetime
import re

# mismas unidades que coop.foja.item (UOM_FISICA)
_UOM = {
    'm2': 'm2', 'm²': 'm2', 'mt2': 'm2',
    'm3': 'm3', 'm³': 'm3', 'mt3': 'm3',
    'ml': 'ml', 'm': 'ml', 'mts': 'ml', 'mt': 'ml',
    'u': 'u', 'un': 'u', 'ud': 'u', 'unidad': 'u', 'c/u': 'u',
    'gl': 'gl', 'global': 'gl',
    'kg': 'kg', 'kgs': 'kg', 'tn': 'kg',
}

_CAB_ITEM = ('ítem', 'item', 'itm', 'nro', 'n°')
_CAB_DESC = ('descripción', 'descripcion', 'detalle', 'tarea', 'designación')
_CAB_UOM = ('u.', 'un.', 'unidad', 'u', 'um')
_CAB_CANT = ('cant.', 'cantidad', 'cant')
_CAB_INC = ('incidencia', 'incidencias')
_CAB_PU = ('precio unitario', 'p. unitario', 'p.unitario', 'unitario')
_CAB_TOT = ('precio total', 'p. total', 'importe', 'monto', 'subtotal', '$',
            'uvis', 'uvi')

_PALABRAS_TOTAL = ('total', 'subtotal', 'suma', 'totales')


def _txt(v):
    return '' if v is None else str(v).strip()


def _norm(v):
    return _txt(v).lower().replace('  ', ' ').strip(' :')


def _num(v):
    """Número tolerante: acepta 1.234,56 y 1,234.56 y '$ 1.000'."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _txt(v).replace('$', '').replace(' ', '').replace('\xa0', '')
    if not s:
        return None
    # si tiene coma y punto, el último separador manda como decimal
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    elif re.fullmatch(r'[1-9]\d{0,2}(\.\d{3})+', s):
        # "1.000" en una planilla argentina es mil, no uno con tres decimales.
        # El patrón exige grupos de exactamente 3 dígitos y que no empiece con
        # cero, así "0.500" sigue leyéndose como medio. Queda un caso
        # ambiguo de verdad —"2.186" escrito como texto podría ser una
        # cantidad con decimales— pero en un cómputo esa celda viene numérica
        # y no pasa por acá.
        s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        return None


def normalizar_uom(v):
    """Devuelve (uom_de_odoo, hubo_que_adivinar)."""
    s = _norm(v).rstrip('.')
    if not s:
        return 'otro', True
    if s in _UOM:
        return _UOM[s], False
    return 'otro', True


def codigo_de_celda(v):
    """Devuelve (codigo, nivel, aviso).

    `nivel` es 0 para un rubro (A, B, I), 1 para `3`/`3.0`, 2 para `3.2`, etc.
    """
    if v is None or _txt(v) == '':
        return '', None, None
    # Excel convierte "5.1" en una fecha: día = ítem, mes = subítem.
    if isinstance(v, (datetime.datetime, datetime.date)):
        cod = '%d.%d' % (v.day, v.month)
        return cod, 2, (
            'La celda vino como fecha %s: Excel interpretó "%s" como una fecha. '
            'Se recuperó el código, pero confirmalo.'
            % (v.strftime('%d/%m/%Y'), cod))
    if isinstance(v, float) and v == int(v):
        return '%d.0' % int(v), 1, None
    if isinstance(v, int):
        return '%d.0' % v, 1, None
    s = _txt(v)
    if isinstance(v, float):
        # 1.10 llega como 1.1: Excel le come el cero final y se pierde el orden
        return s, 2, (
            'El código "%s" vino como número: si en el papel decía "%s0", '
            'Excel le comió el cero final.' % (s, s))
    if re.fullmatch(r'[A-Za-z]{1,3}', s):
        return s.upper(), 0, None
    m = re.fullmatch(r'(\d+)(?:[.\-](\d+))*', s)
    if m:
        return s, s.count('.') + s.count('-') + 1 if '.' in s or '-' in s else 1, None
    return s, None, None


def _fila_encabezado(ws, max_scan=40):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan,
                                         values_only=True), 1):
        celdas = [_norm(c) for c in row]
        tiene_item = any(c in _CAB_ITEM for c in celdas)
        tiene_desc = any(c.startswith(_CAB_DESC) for c in celdas if c)
        if tiene_item and tiene_desc:
            return i, row
    return None, None


def _mapear_columnas(row):
    """Índices de columna por encabezado. Devuelve dict con los que encontró."""
    cols = {}
    for j, c in enumerate(row):
        n = _norm(c)
        if not n:
            continue
        if 'item' not in cols and n in _CAB_ITEM:
            cols['item'] = j
        elif 'desc' not in cols and n.startswith(_CAB_DESC):
            cols['desc'] = j
        elif 'uom' not in cols and n in _CAB_UOM:
            cols['uom'] = j
        elif 'cant' not in cols and n in _CAB_CANT:
            cols['cant'] = j
        elif 'inc' not in cols and n.startswith(_CAB_INC):
            cols['inc'] = j
        elif 'pu' not in cols and n.startswith(_CAB_PU):
            cols['pu'] = j
        elif 'total' not in cols and (n in _CAB_TOT
                                      or n.startswith(_CAB_TOT)):
            cols['total'] = j
    return cols


def parse_foja(fuente, hoja=None):
    """Lee un .xlsx y devuelve la foja clasificada, sin crear nada.

    `fuente` es una ruta o un file-like. `hoja` es el nombre de la solapa; si
    no se pasa, se busca una que se llame como una foja de medición.
    """
    import openpyxl
    wb = openpyxl.load_workbook(fuente, data_only=True, read_only=True)
    if hoja and hoja in wb.sheetnames:
        nombre = hoja
    else:
        candidatas = [s for s in wb.sheetnames
                      if 'foja' in s.lower() or 'medicion' in s.lower()
                      or 'medición' in s.lower() or 'computo' in s.lower()
                      or 'cómputo' in s.lower() or 'presupuesto' in s.lower()]
        nombre = candidatas[0] if candidatas else wb.sheetnames[0]
    ws = wb[nombre]

    avisos = []
    fila_cab, row_cab = _fila_encabezado(ws)
    if not fila_cab:
        return {'hoja': nombre, 'hojas': wb.sheetnames, 'filas': [],
                'niveles': {}, 'columnas': {}, 'encabezado_fila': None,
                'avisos': ['No se encontró la fila de encabezado (se buscó '
                           'una fila con "Ítem" y "Descripción" en las '
                           'primeras 40 filas).']}
    cols = _mapear_columnas(row_cab)
    for req in ('item', 'desc'):
        if req not in cols:
            avisos.append('Falta la columna "%s" en el encabezado.' % req)
    if 'total' not in cols and 'pu' not in cols:
        avisos.append('No se encontró ninguna columna de importe. Sin importe '
                      'no hay incidencia, y sin incidencia el % de avance de '
                      'la obra no significa nada.')

    filas = []
    rubro_actual = ''
    for i, row in enumerate(ws.iter_rows(min_row=fila_cab + 1,
                                         values_only=True), fila_cab + 1):
        def celda(clave):
            j = cols.get(clave)
            return row[j] if j is not None and j < len(row) else None

        crudo_item = celda('item')
        desc = _txt(celda('desc'))
        if _txt(crudo_item) == '' and not desc:
            continue

        cod, nivel, aviso_cod = codigo_de_celda(crudo_item)
        cant = _num(celda('cant'))
        pu = _num(celda('pu'))
        total = _num(celda('total'))
        inc = _num(celda('inc'))
        avisos_fila = [a for a in (aviso_cod,) if a]

        if _norm(desc) in _PALABRAS_TOTAL or _norm(cod) in _PALABRAS_TOTAL:
            clase = 'total'
        elif nivel == 0:
            clase = 'rubro'
            rubro_actual = desc or cod
        elif nivel is None and not cant and not total:
            clase = 'nota'
        else:
            clase = 'item'

        uom, adivinada = normalizar_uom(celda('uom'))
        if clase == 'item' and adivinada and _txt(celda('uom')):
            avisos_fila.append(
                'Unidad "%s" no reconocida: quedó como "Otro".'
                % _txt(celda('uom')))
        if clase == 'item' and not cant:
            avisos_fila.append('Sin cantidad.')
        if clase == 'item' and not (total or pu):
            avisos_fila.append('Sin importe.')

        filas.append({
            'fila': i, 'codigo': cod, 'codigo_crudo': crudo_item,
            'nivel': nivel, 'clase': clase, 'rubro': rubro_actual,
            'descripcion': desc, 'uom': uom, 'uom_crudo': _txt(celda('uom')),
            'cantidad': cant, 'precio_unitario': pu, 'importe': total,
            'incidencia': inc, 'avisos': avisos_fila,
        })

    niveles = {}
    for f in filas:
        if f['clase'] != 'item' or f['nivel'] is None:
            continue
        d = niveles.setdefault(f['nivel'], {'n': 0, 'importe': 0.0})
        d['n'] += 1
        d['importe'] += _importe_de(f) or 0.0

    # El total declarado en la planilla es el único árbitro honesto de qué
    # nivel es "la foja": el nivel cuya suma coincide es el que el que armó el
    # cómputo consideraba la obra. Sugerir, no decidir.
    total_declarado = None
    for f in filas:
        if f['clase'] == 'total':
            cand = _importe_de(f) or f['importe']
            if cand:
                total_declarado = cand
    nivel_sugerido = None
    if total_declarado:
        for n, d in sorted(niveles.items()):
            if d['importe'] and abs(d['importe'] - total_declarado) <= max(
                    1.0, total_declarado * 0.01):
                nivel_sugerido = n
                break
    motivo = 'coincide con el total declarado en la planilla'
    if nivel_sugerido is None and niveles:
        nivel_sugerido = max(niveles.items(),
                             key=lambda kv: kv[1]['importe'])[0]
        motivo = ('la planilla no declara un total en esta solapa; se sugiere '
                  'el nivel de mayor importe, pero es una corazonada, no una '
                  'verificación')

    if len([n for n, d in niveles.items() if d['importe']]) > 1:
        avisos.append(
            'Hay importes en más de un nivel de la numeración (%s). Importar '
            'todos cuenta la obra dos veces: los rubros de nivel 1 ya '
            'contienen a sus subítems. Elegí con qué nivel te quedás.'
            % ', '.join('nivel %d: %d ítems, %s' % (
                n, d['n'], ('%.2f' % d['importe']))
                for n, d in sorted(niveles.items()) if d['importe']))

    con_aviso = sum(1 for f in filas if f['avisos'])
    if con_aviso:
        avisos.append('%d fila(s) con algo raro para revisar a mano.'
                      % con_aviso)

    return {'hoja': nombre, 'hojas': wb.sheetnames,
            'encabezado_fila': fila_cab, 'columnas': cols,
            'filas': filas, 'niveles': niveles,
            'total_declarado': total_declarado,
            'nivel_sugerido': nivel_sugerido,
            'nivel_sugerido_motivo': motivo if niveles else '',
            'avisos': avisos}


def _importe_de(f):
    """Importe de la fila.

    Manda la columna de importe sobre la de precio unitario. En el archivo real
    la columna "Precio Unitario (referencia)" traía 2.47e-08 en varios rubros
    (residuo de una fórmula rota) mientras la columna de importe tenía el
    número bueno: confiar en el unitario habría importado la obra en cero.
    """
    if f.get('importe'):
        return f['importe']
    if f.get('precio_unitario') and f.get('cantidad'):
        return f['precio_unitario'] * f['cantidad']
    return None


def filas_importables(parsed, nivel=None):
    """Las filas que se convertirían en coop.foja.item, con su unitario final.

    Agrega el aviso más peligroso de todos: cuando la planilla trae importe Y
    unitario y no cierran entre sí. Ahí uno de los dos está mal y no se puede
    saber cuál sin mirar el papel.
    """
    out = []
    for f in parsed['filas']:
        if f['clase'] != 'item':
            continue
        if nivel is not None and f['nivel'] != nivel:
            continue
        if not f['cantidad']:
            continue
        importe = _importe_de(f)
        if not importe:
            continue
        pu = importe / f['cantidad']
        avisos = list(f['avisos'])
        declarado = f.get('precio_unitario')
        if declarado and abs(declarado - pu) > max(0.01, abs(pu) * 0.01):
            avisos.append(
                'La planilla dice unitario %.2f pero importe/cantidad da '
                '%.2f. Se usó el importe. Revisalo.' % (declarado, pu))
        out.append(dict(f, precio_unitario_final=pu,
                        importe_final=importe, avisos=avisos))
    return out
